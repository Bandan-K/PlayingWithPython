import re
import xml.etree.ElementTree as ET
import json
import os
import sys
from openpyxl import load_workbook

def parse_android_strings(file_path):
    """Parses Android strings.xml file and returns a dictionary of keys and values."""
    strings = {}
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        for string in root.findall('string'):
            name = string.get('name')
            value = "".join(string.itertext()).strip() if name else None
            if name:
                strings[name] = value
    except Exception as e:
        print(f"Error parsing Android strings: {e}")
    return strings

def parse_ios_strings(file_path):
    """Parses iOS Localizable.strings file and returns a dictionary of keys and values."""
    strings = {}
    # Regex to match "key" = "value"; while handling potential escapes
    pattern = re.compile(r'^"(.+)"\s*=\s*"(.+)"\s*;', re.MULTILINE)
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            matches = pattern.findall(content)
            for key, value in matches:
                strings[key] = value
    except Exception as e:
        print(f"Error parsing iOS strings: {e}")
    return strings

def parse_xlsx_strings(file_path):
    """Parses XLSX file and returns a dictionary of keys and values."""
    strings = {}
    if not os.path.exists(file_path):
        print(f"Warning: {file_path} not found.")
        return strings
        
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        # Header: Key, Value Type, Value, ...
        # Column A is Key (index 0), Column C is Value (index 2)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = row[0]
            value = row[2]
            if key:
                strings[key] = str(value).strip() if value is not None else ""
    except Exception as e:
        print(f"Error parsing XLSX {file_path}: {e}")
    return strings

def compare_dicts(dict_a, dict_b):
    """Compares two string dictionaries and returns differences."""
    keys_a = set(dict_a.keys())
    keys_b = set(dict_b.keys())

    only_in_a = sorted(list(keys_a - keys_b))
    only_in_b = sorted(list(keys_b - keys_a))
    
    common_keys = keys_a & keys_b
    in_both = sorted(list(common_keys))
    
    value_mismatches = []
    for key in common_keys:
        val_a = str(dict_a[key]).strip()
        val_b = str(dict_b[key]).strip()
        if val_a != val_b:
            value_mismatches.append({
                "key": key,
                "a_value": val_a,
                "b_value": val_b
            })
    
    return {
        "only_in_a": only_in_a,
        "only_in_b": only_in_b,
        "in_both": in_both,
        "value_mismatches": sorted(value_mismatches, key=lambda x: x['key'])
    }

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_path = os.path.dirname(script_dir)
    
    excel_dir = os.path.join(base_path, "ExcelData")
    source_dir = os.path.join(base_path, "SourceFiles")
    result_dir = os.path.join(base_path, "Results")
    
    # Input app type
    print("\nSelect App Type (default is customer):")
    print("1. customer")
    print("2. merchant")
    print("3. driver")
    print("4. picker")
    choice = input("Enter number or name [customer]: ").strip().lower()
    
    app_map = {"1": "customer", "2": "merchant", "3": "driver", "4": "picker"}
    app_type = app_map.get(choice, choice if choice else "customer")
    
    android_file = os.path.join(source_dir, "strings.xml")
    ios_file = os.path.join(source_dir, "Localizable.strings")
    
    common_xlsx = os.path.join(excel_dir, "app-common-string.xlsx")
    app_xlsx = os.path.join(excel_dir, f"{app_type}-app-string.xlsx")

    # Output files
    android_ios_output = os.path.join(result_dir, "android_ios_comparison.json")
    app_backend_output = os.path.join(result_dir, f"app_backend_{app_type}_comparison.json")

    # 1. Parsing Apps
    print(f"\n[1/3] Parsing App Strings...")
    android_strings = parse_android_strings(android_file)
    ios_strings = parse_ios_strings(ios_file)
    print(f"Found {len(android_strings)} Android strings and {len(ios_strings)} iOS strings.")
    
    # 2. Parsing Backend (Excel)
    print(f"[2/3] Parsing Backend Strings ({app_type})...")
    common_strings = parse_xlsx_strings(common_xlsx)
    app_specific_strings = parse_xlsx_strings(app_xlsx)
    
    # Merge backend strings (app-specific overrides common if keys collide)
    backend_strings = {**common_strings, **app_specific_strings}
    print(f"Found {len(backend_strings)} total backend strings.")
    
    # 3. Android vs iOS (Keep existing logic)
    print(f"[3/3] Comparing and saving results...")
    android_ios_results = compare_dicts(android_strings, ios_strings)
    android_ios_final = {
        "only_in_android": android_ios_results["only_in_a"],
        "only_in_ios": android_ios_results["only_in_b"],
        "in_both": android_ios_results["in_both"],
        "value_mismatches": [
            {"key": m["key"], "android_value": m["a_value"], "ios_value": m["b_value"]}
            for m in android_ios_results["value_mismatches"]
        ]
    }
    
    with open(android_ios_output, 'w', encoding='utf-8') as f:
        json.dump(android_ios_final, f, indent=4, ensure_ascii=False)

    # 4. Apps vs Backend (Flat value_mismatches)
    all_app_keys = set(android_strings.keys()) | set(ios_strings.keys())
    backend_keys = set(backend_strings.keys())
    
    in_apps_not_in_backend = {
        "android": sorted(list(set(android_strings.keys()) - backend_keys)),
        "ios": sorted(list(set(ios_strings.keys()) - backend_keys))
    }
    
    in_backend_not_in_apps = sorted(list(backend_keys - all_app_keys))
    in_both = sorted(list(all_app_keys & backend_keys))
    
    value_mismatches = []
    
    for key in in_both:
        android_val = android_strings.get(key)
        ios_val = ios_strings.get(key)
        backend_val = backend_strings.get(key)
        
        # Mismatch if any app value doesn't match backend
        mismatch = False
        if android_val is not None and android_val != backend_val:
            mismatch = True
        if ios_val is not None and ios_val != backend_val:
            mismatch = True
            
        if mismatch:
            value_mismatches.append({
                "key": key,
                "android_value": android_val,
                "ios_value": ios_val,
                "backend_value": backend_val
            })
            
    # Sort value mismatches
    value_mismatches = sorted(value_mismatches, key=lambda x: x['key'])
    
    app_backend_final = {
        "in_apps_not_in_backend": in_apps_not_in_backend,
        "in_backend_not_in_apps": in_backend_not_in_apps,
        "in_both": in_both,
        "value_mismatches": value_mismatches
    }
    
    with open(app_backend_output, 'w', encoding='utf-8') as f:
        json.dump(app_backend_final, f, indent=4, ensure_ascii=False)

    print("-" * 40)
    print(f"SUCCESS!")
    print(f"1. Android vs iOS results:   {os.path.basename(android_ios_output)}")
    print(f"2. App vs Backend results:    {os.path.basename(app_backend_output)}")
    print("-" * 40)

if __name__ == "__main__":
    main()
