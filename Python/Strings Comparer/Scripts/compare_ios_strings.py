import re
import os
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

def extract_ios_keys(file_path):
    keys = set()
    pattern = re.compile(r'(?:var|let)\s+\w+(?:\s*:\s*\w+)?\s*=\s*"([^"]+)"')
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found")
        return keys
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            match = pattern.search(line)
            if match:
                keys.add(match.group(1))
    return keys

def extract_localizable_keys(file_path):
    keys = set()
    pattern = re.compile(r'"([^"]+)"\s*=\s*".*?"\s*;', re.DOTALL)
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found")
        return keys
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
        matches = pattern.findall(content)
        for m in matches:
            keys.add(m)
    return keys

def get_excel_mapping(file_path):
    mapping = {}
    if not os.path.exists(file_path):
        print(f"Warning: {file_path} not found")
        return mapping
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) >= 3:
                key = str(row[0]).strip() if row[0] else None
                value = str(row[2]).strip() if row[2] else ""
                if key:
                    mapping[key] = value
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
    return mapping

def get_android_mapping(file_path):
    mapping = {}
    if not os.path.exists(file_path):
        print(f"Warning: {file_path} not found")
        return mapping
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        for string in root.findall('string'):
            name = string.get('name')
            value = "".join(string.itertext()).strip() if name else None
            if name:
                mapping[name] = value
    except Exception as e:
        print(f"Error parsing Android XML: {e}")
    return mapping

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    base_path = os.path.dirname(script_dir)

    # Interactive App Selection
    print("\nSelect App Type (default is customer):")
    print("1. customer")
    print("2. merchant")
    print("3. driver")
    print("4. picker")
    choice = input("Enter number or name [customer]: ").strip().lower()
    
    app_map = {"1": "customer", "2": "merchant", "3": "driver", "4": "picker"}
    app_type = app_map.get(choice, choice if choice else "customer")
    
    excel_dir = os.path.join(base_path, 'ExcelData')
    source_dir = os.path.join(base_path, 'SourceFiles')
    result_dir = os.path.join(base_path, 'Results')
    
    ios_file = os.path.join(source_dir, f'{app_type}/iOSFile')
    strings_file = os.path.join(source_dir, f'{app_type}/Localizable.strings')
    android_file = os.path.join(source_dir, f'{app_type}/strings.xml')
    output_file = os.path.join(result_dir, 'missing_keys.txt')

    print("Re-parsing source files...")
    ios_keys = extract_ios_keys(ios_file)
    strings_keys = extract_localizable_keys(strings_file)
    missing = sorted(list(ios_keys - strings_keys))
    
    print(f"Found {len(missing)} missing keys in Localizable.strings")

    common_xlsx = os.path.join(excel_dir, 'app-common-string.xlsx')
    app_xlsx = os.path.join(excel_dir, f'{app_type}-app-string.xlsx')

    print(f"Loading Excel mapping ({app_type})...")
    common_map = get_excel_mapping(common_xlsx)
    app_map_data = get_excel_mapping(app_xlsx)
    excel_mapping = {**common_map, **app_map_data}

    print(f"Loading Android mapping (strings.xml)...")
    android_mapping = get_android_mapping(android_file)

    with open(output_file, 'w', encoding='utf-8') as f:
        for key in missing:
            # 1. Check Excel
            value = excel_mapping.get(key)
            # 2. Fallback to Android XML
            if not value and key in android_mapping:
                value = android_mapping[key]
            
            # Final fallback to empty string
            if value is None:
                value = ""
                
            f.write(f'"{key}" = "{value}";\n')

    print(f"\nDone! Updated results saved to {output_file}")

if __name__ == "__main__":
    main()
